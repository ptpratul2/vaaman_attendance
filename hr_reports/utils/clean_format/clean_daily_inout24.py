# clean_daily_inout24.py
import os
import pandas as pd
import frappe
from datetime import datetime, timedelta, time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def clean_daily_inout24(input_path: str, output_path: str, company: str = None, branch: str = None) -> pd.DataFrame:
    print("=" * 80)
    print("[clean_daily_inout24] Starting")
    print(f"[clean_daily_inout24] Input: {input_path}")
    print(f"[clean_daily_inout24] Output: {output_path}")
    print(f"[clean_daily_inout24] Company: {company}")
    print(f"[clean_daily_inout24] Branch: {branch}")
    print("=" * 80)

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Load the file
    df_raw = pd.read_excel(input_path, engine="openpyxl")
    print(f"[clean_daily_inout24] Loaded raw DataFrame shape: {df_raw.shape}")

    # Required columns from raw report
    required_cols = ["Name", "Gate Pass", "Date", "Intime", "Outtime", "GROSSHOURS", "Shift"]
    missing = [c for c in required_cols if c not in df_raw.columns]
    if missing:
        raise ValueError(f"Missing required columns in input: {missing}")

    # ------------------------------------------------------
    # Function: format "Extra/Less Hours" → decimal-like str
    # ------------------------------------------------------
    def format_extra_less(v):
        if pd.isna(v):
            return ""

        # Case 1: Excel gave us a datetime.time or Timestamp
        if hasattr(v, "hour") and hasattr(v, "minute"):
            h = v.hour
            m = v.minute
            s = v.second
            # If there are seconds → keep them as .ss, else only hh.mm
            if s:
                return f"{h}.{s:02d}"
            else:
                return f"{h}.{m:02d}"

        # Case 2: It's already a string like -1:-22: or 4:30
        s = str(v).strip()
        if not s:
            return ""

        # Replace last ":" with "."
        if ":" in s:
            parts = s.split(":")
            if len(parts) >= 2:
                return ":".join(parts[:-1]) + "." + parts[-1]

        return s

    def parse_time_to_hours(time_str):
        """Convert time string (HH:MM or HH:MM:SS) to decimal hours"""
        if not time_str or pd.isna(time_str):
            return 0.0
        try:
            time_str = str(time_str).strip()
            parts = time_str.split(":")
            hours = int(parts[0])
            minutes = int(parts[1]) if len(parts) > 1 else 0
            seconds = int(parts[2]) if len(parts) > 2 else 0
            return hours + minutes/60 + seconds/3600
        except Exception:
            return 0.0

    def detect_shift_from_checkin(checkin_time_str, att_date, find_nearest=False):
        """
        Detect shift based on check-in time
        Shift definitions (punch-in time windows):
        - A shift: punch between 5-7 (05:00 to 07:00)
        - G shift: punch between 8-10 (08:00 to 10:00)
        - B shift: punch between 13-15 (13:00 to 15:00)
        - C shift: punch between 21-23 (21:00 to 23:00)

        If find_nearest=True, returns the nearest shift when outside windows
        """
        if not checkin_time_str or not att_date:
            return None

        try:
            # Parse check-in time
            checkin_parts = str(checkin_time_str).strip().split(":")
            checkin_hour = int(checkin_parts[0])
            checkin_min = int(checkin_parts[1]) if len(checkin_parts) > 1 else 0
            checkin_time = time(checkin_hour, checkin_min)

            # Parse date
            date_obj = pd.to_datetime(att_date).date()
            checkin_datetime = datetime.combine(date_obj, checkin_time)

            # Shift definitions based on punch-in time windows:
            # A shift: punch between 5-7 (05:00 to 07:00)
            # G shift: punch between 8-10 (08:00 to 10:00)
            # B shift: punch between 13-15 (13:00 to 15:00)
            # C shift: punch between 21-23 (21:00 to 23:00)

            a_start = datetime.combine(date_obj, time(5, 0))
            a_end = datetime.combine(date_obj, time(7, 0))

            g_start = datetime.combine(date_obj, time(8, 0))
            g_end = datetime.combine(date_obj, time(10, 0))

            b_start = datetime.combine(date_obj, time(13, 0))
            b_end = datetime.combine(date_obj, time(15, 0))

            c_start = datetime.combine(date_obj, time(21, 0))
            c_end = datetime.combine(date_obj, time(23, 0))

            # Check which shift the check-in time falls into
            if a_start <= checkin_datetime <= a_end:
                return "A"
            elif g_start <= checkin_datetime <= g_end:
                return "G"
            elif b_start <= checkin_datetime <= b_end:
                return "B"
            elif c_start <= checkin_datetime <= c_end:
                return "C"
            else:
                # If outside all windows
                if find_nearest:
                    # Find nearest shift based on time
                    # A center: 6:00, G center: 9:00, B center: 14:00, C center: 22:00
                    a_center = datetime.combine(date_obj, time(6, 0))
                    g_center = datetime.combine(date_obj, time(9, 0))
                    b_center = datetime.combine(date_obj, time(14, 0))
                    c_center = datetime.combine(date_obj, time(22, 0))

                    # Calculate distances
                    distances = {
                        "A": abs((checkin_datetime - a_center).total_seconds()),
                        "G": abs((checkin_datetime - g_center).total_seconds()),
                        "B": abs((checkin_datetime - b_center).total_seconds()),
                        "C": abs((checkin_datetime - c_center).total_seconds())
                    }

                    # Return nearest shift
                    nearest = min(distances, key=distances.get)
                    return nearest
                else:
                    return None

        except Exception as e:
            print(f"[clean_daily_inout24] Error detecting shift: {e}")
            return None

    def parse_to_datetime(time_value, att_date):
        """
        Parse time value to datetime object for ERPNext
        Handles both datetime objects and time strings
        """
        if not time_value or not att_date:
            return None

        try:
            # If it's already a datetime object
            if isinstance(time_value, datetime):
                return time_value
            
            # If it's a pandas Timestamp
            if isinstance(time_value, pd.Timestamp):
                return time_value.to_pydatetime()
            
            # Parse date
            date_obj = pd.to_datetime(att_date).date()

            # Parse time string
            time_str = str(time_value).strip()
            # Handle datetime string format (e.g., "2024-01-15 09:30:00")
            if " " in time_str:
                return pd.to_datetime(time_str).to_pydatetime()
            
            # Handle time-only string (HH:MM or HH:MM:SS)
            time_parts = time_str.split(":")
            in_hour = int(time_parts[0])
            in_min = int(time_parts[1]) if len(time_parts) > 1 else 0
            in_sec = int(time_parts[2]) if len(time_parts) > 2 else 0

            # Create datetime object
            return datetime.combine(date_obj, time(in_hour, in_min, in_sec))
        except Exception as e:
            print(f"[clean_daily_inout24] Error parsing datetime: {e}, value: {time_value}")
            return None

    def calculate_working_hours(intime, outtime, att_date):
        """Calculate working hours from intime and outtime (datetime objects)"""
        if not intime or not outtime or not att_date:
            return None, 0.0

        try:
            # Parse to datetime objects
            in_dt = parse_to_datetime(intime, att_date)
            out_dt = parse_to_datetime(outtime, att_date)

            if not in_dt or not out_dt:
                return None, 0.0

            # If outtime is earlier than intime, assume it's next day
            if out_dt < in_dt:
                out_dt += timedelta(days=1)

            # Calculate difference
            diff = out_dt - in_dt
            total_seconds = diff.total_seconds()
            hours = total_seconds / 3600

            # Format as decimal (e.g., 16.00)
            work_hrs_float = round(hours, 2)

            return work_hrs_float, hours
        except Exception as e:
            print(f"[clean_daily_inout24] Error calculating hours: {e}")
            return None, 0.0

    records = []

    for _, row in df_raw.iterrows():
        gate_pass = str(row.get("Gate Pass")).strip() if pd.notna(row.get("Gate Pass")) else None
        emp_name = str(row.get("Name")).strip() if pd.notna(row.get("Name")) else None
        att_date = pd.to_datetime(row.get("Date")).strftime("%Y-%m-%d") if pd.notna(row.get("Date")) else None
        intime_raw = row.get("Intime") if pd.notna(row.get("Intime")) else None
        outtime_raw = row.get("Outtime") if pd.notna(row.get("Outtime")) else None
        gross_hours = str(row.get("GROSSHOURS")).strip() if pd.notna(row.get("GROSSHOURS")) else None
        shift = str(row.get("Shift")).strip() if pd.notna(row.get("Shift")) and str(row.get("Shift")).strip() else None
        over_time = format_extra_less(row.get("Extra/Less Hours"))

        # ------------------------
        # Parse intime/outtime to datetime objects for ERPNext
        # ------------------------
        intime_dt = parse_to_datetime(intime_raw, att_date) if intime_raw else None
        outtime_dt = parse_to_datetime(outtime_raw, att_date) if outtime_raw else None

        # Format datetime for ERPNext (YYYY-MM-DD HH:MM:SS)
        intime_str = intime_dt.strftime("%Y-%m-%d %H:%M:%S") if intime_dt else ""
        outtime_str = outtime_dt.strftime("%Y-%m-%d %H:%M:%S") if outtime_dt else ""

        # ------------------------
        # Store original shift from Excel for fallback
        # ------------------------
        original_shift = shift

        # ------------------------
        # Handle "N" shift - auto-detect actual shift from punch time
        # ------------------------
        if shift and shift.upper() == "N":
            print(f"[clean_daily_inout24] Shift 'N' detected for {emp_name} - will auto-detect shift from punch time")
            if intime_dt:
                # Try to detect shift from punch time
                detected_shift = detect_shift_from_checkin(intime_raw, att_date, find_nearest=False)
                if detected_shift:
                    shift = detected_shift
                    print(f"[clean_daily_inout24] Auto-detected shift '{shift}' for {emp_name} with check-in {intime_str}")
                else:
                    # If outside windows, find nearest shift
                    nearest_shift = detect_shift_from_checkin(intime_raw, att_date, find_nearest=True)
                    if nearest_shift:
                        shift = nearest_shift
                        print(f"[clean_daily_inout24] Assigned nearest shift '{shift}' for {emp_name} (punch outside windows)")
                    else:
                        shift = ""
            else:
                shift = ""  # No punch time, clear shift

        # ------------------------
        # Auto-detect shift if blank
        # ------------------------
        elif not shift and intime_dt:
            detected_shift = detect_shift_from_checkin(intime_raw, att_date, find_nearest=False)
            if detected_shift:
                shift = detected_shift
                print(f"[clean_daily_inout24] Auto-detected shift '{shift}' for {emp_name} with check-in {intime_str}")
            else:
                # If outside windows, keep original shift from Excel (if not "N")
                if original_shift and original_shift.upper() != "N":
                    shift = original_shift
                    print(f"[clean_daily_inout24] Using original shift '{shift}' from Excel (punch outside detection windows)")
                else:
                    shift = ""

        # ------------------------
        # Map Gate Pass → Employee ID
        # ------------------------
        employee_id = None
        try:
            emp_doc = frappe.get_doc("Employee", {"attendance_device_id": gate_pass})
            employee_id = emp_doc.name
        except Exception:
            print(f"[clean_daily_inout24] WARNING: Employee not found for Gate Pass {gate_pass}")

        # ------------------------
        # Calculate Working Hours and Determine Status
        # ------------------------
        status = "Absent"
        working_hours_float = 0.0  # Will be calculated from Intime/Outtime

        # If both Intime and Outtime are present, calculate working hours
        if intime_dt and outtime_dt:
            calc_work_hrs, total_hours = calculate_working_hours(intime_dt, outtime_dt, att_date)

            if calc_work_hrs is not None:
                working_hours_float = calc_work_hrs

                # Determine status based on working hours
                if total_hours >= 7:
                    status = "Present"
                elif total_hours >= 4.5:  # 4:30 = 4.5 hours
                    status = "Half Day"
                else:
                    status = "Absent"
        # If Intime or Outtime is missing, mark as Absent with no working hours
        else:
            status = "Absent"
            working_hours_float = 0.0

        # ------------------------
        # Final rule: If working hours = 0 or status = "Absent", clear shift
        # ------------------------
        if working_hours_float == 0.0 or status == "Absent":
            if shift:
                print(f"[clean_daily_inout24] Clearing shift '{shift}' for {emp_name} (Working Hours: {working_hours_float}, Status: {status})")
            shift = ""

        # ------------------------
        # Don't show working hours if 0 or negative
        # ------------------------
        display_working_hours = working_hours_float if working_hours_float > 0 else ""

        # ------------------------
        # Don't show overtime if 0 or negative
        # Fix malformed values like "9.-7" → "9.00"
        # ------------------------
        display_overtime = ""
        if over_time:
            try:
                overtime_str = str(over_time).strip()
                if overtime_str and overtime_str not in ["0", "0.0", "0.00", ""]:
                    # If minutes part is negative (like "9.-7"), treat minutes as 0
                    if ".-" in overtime_str:
                        hours_part = overtime_str.split(".-")[0]
                        overtime_str = f"{hours_part}.00"

                    # Now check if positive
                    if "." in overtime_str or ":" in overtime_str:
                        parts = overtime_str.replace(":", ".").split(".")
                        hours = int(parts[0]) if parts[0] and parts[0] not in ['-', ''] else 0
                        minutes = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
                        # Show only if positive
                        if hours > 0 or minutes > 0:
                            display_overtime = f"{hours}.{minutes:02d}" if minutes else str(hours)
                    elif float(overtime_str) > 0:
                        display_overtime = overtime_str
            except Exception:
                display_overtime = ""

        rec = {
            "Attendance Date": att_date,
            "Employee": employee_id if employee_id else "",
            "Employee Name": emp_name,
            "Status": status,
            "In Time": intime_str,  # Datetime format for ERPNext
            "Out Time": outtime_str,  # Datetime format for ERPNext
            "Company": company if company else "",
            "Branch": branch if branch else "",
            "Working Hours": display_working_hours,
            "Shift": shift if shift else "",
            "Over Time": display_overtime
        }
        records.append(rec)

    df_final = pd.DataFrame.from_records(records)

    # Drop rows without employee/date
    df_final = df_final.dropna(subset=["Attendance Date", "Employee"], how="any")
    print(f"[clean_daily_inout24] Built final DataFrame with {len(df_final)} rows")

    if df_final.empty:
        raise ValueError("No attendance records parsed from Daily In-Out report.")

    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    df_final.to_excel(output_path, index=False)
    print(f"[clean_daily_inout24] Saved output to: {output_path}")

    # Apply Excel formatting: black out cells where working hours < 0
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        # Find the Working Hours column index
        working_hours_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Working Hours":
                working_hours_col = col_idx
                break

        if working_hours_col:
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

            # Iterate through rows (skip header)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=working_hours_col)
                if cell.value is not None:
                    try:
                        value = float(cell.value)
                        if value < 0:
                            cell.fill = black_fill
                            print(f"[clean_daily_inout24] Blacked out cell at row {row_idx} (Working Hours: {value})")
                    except (ValueError, TypeError):
                        pass  # Skip non-numeric values

            wb.save(output_path)
            print(f"[clean_daily_inout24] Applied cell formatting for negative working hours")
    except Exception as e:
        print(f"[clean_daily_inout24] Warning: Could not apply Excel formatting: {e}")

    print("[clean_daily_inout24] Done ✅")

    return df_final
