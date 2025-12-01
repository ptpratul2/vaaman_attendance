# clean_daily_inout7_2.py
# Cleaning script for VAAMAN ENGINEERS Daily Attendance Report format
# Format: Separate Date In/Time In and Date Out/Time Out columns

import os
import pandas as pd
import frappe
from datetime import datetime, timedelta, time
from openpyxl import load_workbook

def clean_daily_inout7_2(input_path: str, output_path: str, company: str = None, branch: str = None) -> pd.DataFrame:
    """
    Clean VAAMAN ENGINEERS Daily Attendance Report Excel format

    Input Excel Structure:
    - Row 0: Company name header
    - Row 1: Blank
    - Row 2: Column headers (Emp Code, Employee Name, etc.)
    - Row 3+: Data rows

    Columns:
    - Emp Code (attendance_device_id)
    - Employee Name
    - Fathers Name
    - Work Order No
    - Designation
    - Department
    - Date In (DD-MM-YYYY)
    - Time In (HH:MM:SS)
    - Date Out (DD-MM-YYYY) - may be missing
    - Time Out (HH:MM:SS) - may be missing
    """

    print("=" * 80)
    print("[clean_daily_inout_vaaman] Starting VAAMAN format cleaning")
    print(f"[clean_daily_inout_vaaman] Input: {input_path}")
    print(f"[clean_daily_inout_vaaman] Output: {output_path}")
    print(f"[clean_daily_inout_vaaman] Company: {company}")
    print(f"[clean_daily_inout_vaaman] Branch: {branch}")
    print("=" * 80)

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Load Excel file - skip first 3 rows (company name + blank + headers)
    # Manually assign column names to avoid issues with merged cells
    df_raw = pd.read_excel(
        input_path,
        engine="openpyxl",
        skiprows=3,
        names=[
            'Emp Code', 'Employee Name', 'Fathers Name', 'Work Order No',
            'Designation', 'Department', 'Date In', 'Time In', 'Date Out', 'Time Out'
        ]
    )

    print(f"[clean_daily_inout_vaaman] Loaded raw DataFrame shape: {df_raw.shape}")

    # Remove rows with missing Emp Code (blank rows)
    df_raw = df_raw[df_raw["Emp Code"].notna()]
    print(f"[clean_daily_inout_vaaman] After removing blank rows: {df_raw.shape}")

    def parse_datetime(date_val, time_val):
        """
        Combine separate date and time values into datetime object
        Handles various date formats (DD-MM-YYYY, datetime objects, etc.)
        """
        if pd.isna(date_val) or pd.isna(time_val):
            return None

        try:
            # Parse date
            if isinstance(date_val, datetime):
                date_obj = date_val.date()
            elif isinstance(date_val, pd.Timestamp):
                date_obj = date_val.date()
            else:
                # Handle string format DD-MM-YYYY
                date_str = str(date_val).strip()
                date_obj = pd.to_datetime(date_str, format='%d-%m-%Y', errors='coerce')
                if pd.isna(date_obj):
                    date_obj = pd.to_datetime(date_str, errors='coerce')
                if pd.notna(date_obj):
                    date_obj = date_obj.date()
                else:
                    return None

            # Parse time
            if isinstance(time_val, time):
                time_obj = time_val
            elif isinstance(time_val, datetime):
                time_obj = time_val.time()
            elif isinstance(time_val, pd.Timestamp):
                time_obj = time_val.time()
            else:
                # Handle string format HH:MM:SS
                time_str = str(time_val).strip()
                time_parts = time_str.split(":")
                if len(time_parts) >= 2:
                    hour = int(time_parts[0])
                    minute = int(time_parts[1])
                    second = int(time_parts[2]) if len(time_parts) > 2 else 0
                    time_obj = time(hour, minute, second)
                else:
                    return None

            # Combine date and time
            return datetime.combine(date_obj, time_obj)

        except Exception as e:
            print(f"[clean_daily_inout_vaaman] Error parsing datetime: {e}, date={date_val}, time={time_val}")
            return None

    def calculate_working_hours(checkin_dt, checkout_dt):
        """
        Calculate working hours from check-in and check-out datetime objects
        If checkout < checkin, assume next day checkout
        Returns: (working_hours_float, total_hours_for_status_check)
        """
        if not checkin_dt or not checkout_dt:
            return 0.0, 0.0

        try:
            # If checkout is earlier than checkin, assume it's next day
            if checkout_dt < checkin_dt:
                checkout_dt += timedelta(days=1)

            # Calculate difference
            diff = checkout_dt - checkin_dt
            total_seconds = diff.total_seconds()
            hours = total_seconds / 3600

            return round(hours, 2), hours

        except Exception as e:
            print(f"[clean_daily_inout_vaaman] Error calculating hours: {e}")
            return 0.0, 0.0

    def detect_shift(checkin_dt):
        """
        Auto-detect shift based on check-in time

        Shift definitions:
        - A shift: 6 AM to 2 PM (check-in 5:30 AM to 2 PM)
        - B shift: 2 PM to 10 PM (check-in 1:30 PM to 10 PM)
        - C shift: 10 PM to 6 AM (check-in 9:30 PM to 6 AM next day) - night shift
        - G shift: 9 AM to 5:30 PM (check-in 8:30 AM to 5:30 PM) - general shift
        """
        if not checkin_dt:
            return None

        try:
            checkin_time = checkin_dt.time()

            # Define shift time ranges (with 30-minute buffer before start)
            c_start = time(21, 30)  # 9:30 PM
            c_end = time(5, 30)     # 5:30 AM next day
            g_start = time(8, 30)   # 8:30 AM
            g_end = time(17, 30)    # 5:30 PM
            a_start = time(5, 30)   # 5:30 AM
            a_end = time(14, 0)     # 2:00 PM
            b_start = time(13, 30)  # 1:30 PM
            b_end = time(22, 0)     # 10:00 PM

            # Priority: C (night) > G (general) > A > B

            # C shift spans midnight (9:30 PM to 5:30 AM)
            if checkin_time >= c_start or checkin_time < c_end:
                return "C"
            # G shift (8:30 AM to 5:30 PM) - check before A/B to handle overlap
            elif g_start <= checkin_time <= g_end:
                return "G"
            # A shift (5:30 AM to 2 PM)
            elif a_start <= checkin_time <= a_end:
                return "A"
            # B shift (1:30 PM to 10 PM)
            elif b_start <= checkin_time <= b_end:
                return "B"
            else:
                # No shift matches - return None (blank cell)
                return None

        except Exception as e:
            print(f"[clean_daily_inout_vaaman] Error detecting shift: {e}")
            return None

    def determine_status(working_hours):
        """
        Determine attendance status based on working hours
        - Present: >= 7 hours
        - Half Day: >= 4.5 hours and < 7 hours
        - Absent: < 4.5 hours or missing check-in/out
        """
        if working_hours >= 7.0:
            return "Present"
        elif working_hours >= 4.5:
            return "Half Day"
        else:
            return "Absent"

    # First pass: Group punches by employee and date
    from collections import defaultdict
    grouped_punches = defaultdict(list)

    for idx, row in df_raw.iterrows():
        # Extract raw values
        emp_code = str(int(row["Emp Code"])) if pd.notna(row["Emp Code"]) else None
        emp_name = str(row["Employee Name"]).strip() if pd.notna(row["Employee Name"]) else None
        date_in = row["Date In"]
        time_in = row["Time In"]
        date_out = row["Date Out"]
        time_out = row["Time Out"]

        # Skip if no emp code or emp name
        if not emp_code or not emp_name:
            continue

        # Parse check-in datetime
        checkin_dt = parse_datetime(date_in, time_in)

        # Parse check-out datetime (may be missing)
        checkout_dt = parse_datetime(date_out, time_out) if pd.notna(date_out) and pd.notna(time_out) else None

        # Skip if no valid check-in
        if not checkin_dt:
            continue

        # Get attendance date from check-in
        att_date = checkin_dt.strftime("%Y-%m-%d")

        # Group by (emp_code, emp_name, att_date)
        key = (emp_code, emp_name, att_date)
        grouped_punches[key].append((checkin_dt, checkout_dt))

    # Second pass: Process grouped punches - sum working hours
    records = []
    incomplete_count = 0
    employee_not_found_count = 0
    merged_count = 0

    for (emp_code, emp_name, att_date), punches in grouped_punches.items():
        if len(punches) > 1:
            merged_count += 1
            print(f"[clean_daily_inout_vaaman] MERGED: {emp_name} on {att_date} - {len(punches)} punches")

        # Calculate working hours using FIRST In Time → LAST Out Time approach
        # This handles cases where employees forget to punch out during shifts
        first_checkin = None
        last_checkout = None
        punch_count = 0

        for checkin_dt, checkout_dt in punches:
            # Track FIRST check-in (earliest) from ALL punches
            if checkin_dt:
                if first_checkin is None or checkin_dt < first_checkin:
                    first_checkin = checkin_dt
                punch_count += 1

            # Track LAST check-out (latest) from ALL punches
            if checkout_dt:
                if last_checkout is None or checkout_dt > last_checkout:
                    last_checkout = checkout_dt

        # Calculate working hours from FIRST In → LAST Out
        if first_checkin and last_checkout:
            # Handle night shift crossing midnight
            checkout_dt_calc = last_checkout
            if checkout_dt_calc < first_checkin:
                checkout_dt_calc = checkout_dt_calc + timedelta(days=1)
                print(f"[clean_daily_inout_vaaman]   Night shift detected: adjusted checkout to next day")

            diff = checkout_dt_calc - first_checkin
            total_working_hours = diff.total_seconds() / 3600
            working_hours_float = round(total_working_hours, 2)

            print(f"[clean_daily_inout_vaaman]   First In: {first_checkin.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"[clean_daily_inout_vaaman]   Last Out: {checkout_dt_calc.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"[clean_daily_inout_vaaman]   Total Working Hours: {working_hours_float}h (from first in to last out)")
        else:
            working_hours_float = 0.0
            print(f"[clean_daily_inout_vaaman]   ⚠️ Incomplete attendance - missing checkout")

        # Determine status based on total working hours
        if working_hours_float >= 7.0:
            status = "Present"
        elif working_hours_float >= 4.5:
            status = "Half Day"
        else:
            status = "Absent"

        # If no valid checkout, mark as incomplete
        if not first_checkin or not last_checkout:
            incomplete_count += 1
            status = "Absent"

        # Auto-detect shift from first check-in
        shift = detect_shift(first_checkin) if first_checkin else None

        # Calculate Over Time (hours worked beyond 8 hours)
        overtime_hours = round(working_hours_float - 8.0, 2)
        print(f"[clean_daily_inout_vaaman]   Overtime: {overtime_hours}h")

        # Map Emp Code to Employee ID
        employee_id = None
        try:
            emp_doc = frappe.get_doc("Employee", {"attendance_device_id": emp_code})
            employee_id = emp_doc.name
        except Exception:
            print(f"[clean_daily_inout_vaaman] WARNING: Employee not found for Emp Code {emp_code}")
            employee_not_found_count += 1

        # Format datetime strings for ERPNext (YYYY-MM-DD HH:MM:SS) - 24-hour format
        checkin_str = first_checkin.strftime("%Y-%m-%d %H:%M:%S") if first_checkin else ""
        checkout_str = last_checkout.strftime("%Y-%m-%d %H:%M:%S") if last_checkout else ""

        # Build record
        rec = {
            "Attendance Date": att_date,
            "Employee": employee_id if employee_id else "",
            "Employee Name": emp_name,
            "Status": status,
            "In Time": checkin_str,
            "Out Time": checkout_str,
            "Company": company if company else "",
            "Branch": branch if branch else "",
            "Working Hours": working_hours_float,
            "Shift": shift if shift else "",
            "Over Time": overtime_hours,
        }
        records.append(rec)

    print(f"[clean_daily_inout_vaaman] Processed {len(records)} attendance records")
    print(f"[clean_daily_inout_vaaman] Merged multi-punch records: {merged_count}")
    print(f"[clean_daily_inout_vaaman] Incomplete records (missing check-out): {incomplete_count}")
    print(f"[clean_daily_inout_vaaman] Employees not found in system: {employee_not_found_count}")

    # Create final DataFrame
    df_final = pd.DataFrame.from_records(records)

    # Drop rows without employee ID or attendance date
    initial_count = len(df_final)
    df_final = df_final.dropna(subset=["Attendance Date", "Employee"], how="any")
    dropped_count = initial_count - len(df_final)

    if dropped_count > 0:
        print(f"[clean_daily_inout_vaaman] WARNING: Dropped {dropped_count} rows due to missing Employee ID")

    print(f"[clean_daily_inout_vaaman] Final DataFrame shape: {df_final.shape}")

    if df_final.empty:
        raise ValueError("No valid attendance records found after processing.")

    # Create output directory if needed
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    # Save to Excel
    df_final.to_excel(output_path, index=False, engine='openpyxl')

    # Make Over Time cell blank (empty) if value < 1.0
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        # Find Over Time column index
        overtime_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Over Time":
                overtime_col_idx = idx
                break

        if overtime_col_idx:
            # Iterate through data rows (skip header row 1)
            blank_cells_count = 0
            for row_idx in range(2, ws.max_row + 1):
                overtime_cell = ws.cell(row=row_idx, column=overtime_col_idx)

                # Check if Over Time value is less than 1.0
                try:
                    overtime_value = float(overtime_cell.value) if overtime_cell.value is not None else 0.0

                    if overtime_value < 1.0:
                        # Make the cell blank (empty)
                        overtime_cell.value = None
                        blank_cells_count += 1
                except (ValueError, TypeError):
                    # Skip if overtime value is not a number
                    pass

            # Save the workbook with blanked cells
            wb.save(output_path)
            print(f"[clean_daily_inout_vaaman] Made {blank_cells_count} overtime cells blank (overtime < 1.0)")

    except Exception as e:
        print(f"[clean_daily_inout_vaaman] Warning: Could not blank cells: {e}")

    print(f"[clean_daily_inout_vaaman] Saved cleaned output to: {output_path}")
    print("[clean_daily_inout_vaaman] Done ✅")
    print("=" * 80)

    return df_final
